#!/usr/bin/env python3
"""Generate standalone deliverables for SAT 2000 (all 40 days, 2000 words).

Outputs (next to this script):
  - SAT_Vocab_2000_All.xlsx   (openpyxl) — "All" sheet + 40 per-day sheets
  - SAT_Vocab_2000_All.pdf    (reportlab) — full list grouped by day

Source of truth is data/days.js. We read it via Node into JSON so the offline
files always match the website data. Run:  python3 exports/build_exports.py
"""
import json
import os
import subprocess
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
COLS = ["Day", "Theme (EN)", "Theme (KO)", "Word", "POS", "IPA", "English", "Korean", "Example"]
COLS_DAY = ["Word", "POS", "IPA", "English", "Korean", "Example"]
ACCENT = "1FC85C"


def load_days():
    """Read data/days.js via Node so the JS file stays the single source."""
    script = (
        "global.window={};require('%s');"
        "process.stdout.write(JSON.stringify(window.DAYS.filter(d=>d.words.length>0)));"
        % os.path.join(ROOT, "data", "days.js").replace("\\", "/")
    )
    out = subprocess.check_output(["node", "-e", script])
    return json.loads(out)


def build_xlsx(days, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor=ACCENT)
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="E2E2E4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── "All" sheet: Day, Theme (EN), Theme (KO), Word, POS, IPA, English, Korean, Example ──
    ws_all = wb.create_sheet(title="All")
    ws_all.append(COLS)
    all_widths = [5, 22, 18, 16, 6, 24, 34, 18, 70]
    for c, w in enumerate(all_widths, start=1):
        ws_all.column_dimensions[chr(64 + c)].width = w
        cell = ws_all.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        cell.border = border
    ws_all.freeze_panes = "A2"

    for d in days:
        for word in d["words"]:
            ws_all.append([
                d["day"], d["theme"], d["theme_ko"],
                word["w"], word["pos"], word["ipa"],
                word["en"], word["ko"], word["ex"]
            ])

    for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row, max_col=9):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 9))
            cell.border = border

    # ── Per-day sheets ──
    day_widths = [16, 6, 24, 34, 18, 70]
    for d in days:
        ws = wb.create_sheet(title="Day %02d" % d["day"])
        ws.append(COLS_DAY)
        for c, w in enumerate(day_widths, start=1):
            ws.column_dimensions[chr(64 + c)].width = w
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        for word in d["words"]:
            ws.append([word["w"], word["pos"], word["ipa"],
                       word["en"], word["ko"], word["ex"]])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 6))
                cell.border = border
        ws.freeze_panes = "A2"

    wb.save(path)
    return path


def build_pdf(days, path):
    from datetime import date
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, PageBreak)

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("cell", parent=styles["BodyText"],
                                fontSize=7.5, leading=9.5)
    mono_style = ParagraphStyle("mono", parent=cell_style, fontName="Courier")
    head_style = ParagraphStyle("head", parent=styles["Heading2"],
                                fontSize=13, spaceAfter=6,
                                textColor=colors.HexColor("#111111"))
    printed_style = ParagraphStyle("printed", parent=styles["BodyText"],
                                   fontSize=9, textColor=colors.HexColor("#969696"),
                                   spaceAfter=8)
    accent = colors.HexColor("#" + ACCENT)

    doc = SimpleDocTemplate(
        path, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title="SAT 2000 — All 2000 Words")
    story = []
    col_w = [70, 32, 95, 150, 78, 290]  # points; sums under landscape A4 usable width

    # Cover printed date — inserted once before all day sections
    story.append(Paragraph("Printed: " + date.today().isoformat(), printed_style))

    for di, d in enumerate(days):
        story.append(Paragraph(
            "SAT 2000 &mdash; Day %02d &middot; %s (%s)" % (
                d["day"], d["theme"], d["theme_ko"]),
            head_style))
        data = [[Paragraph("<b>%s</b>" % c, cell_style) for c in COLS_DAY]]
        for w in d["words"]:
            data.append([
                Paragraph(w["w"], mono_style),
                Paragraph(w["pos"], cell_style),
                Paragraph(w["ipa"], mono_style),
                Paragraph(w["en"], cell_style),
                Paragraph(w["ko"], cell_style),
                Paragraph(w["ex"], cell_style),
            ])
        tbl = Table(data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), accent),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E2E4")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F6F6F6")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(tbl)
        if di < len(days) - 1:
            story.append(PageBreak())

    doc.build(story)
    return path


def main():
    days = load_days()
    total = sum(len(d["words"]) for d in days)
    print("Days loaded: %d | total words: %d" % (len(days), total))

    xlsx_path = os.path.join(HERE, "SAT_Vocab_2000_All.xlsx")
    pdf_path = os.path.join(HERE, "SAT_Vocab_2000_All.pdf")

    xlsx = build_xlsx(days, xlsx_path)
    print("XLSX: %s (%d bytes)" % (xlsx, os.path.getsize(xlsx)))

    pdf = build_pdf(days, pdf_path)
    print("PDF : %s (%d bytes)" % (pdf, os.path.getsize(pdf)))


if __name__ == "__main__":
    main()
